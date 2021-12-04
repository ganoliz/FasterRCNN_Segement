# -*- coding: utf-8 -*-
"""
Created on Mon Nov 22 14:43:06 2021

@author: Nick
"""

from PyQt5 import QtWidgets,QtCore
from Spine_GUI2 import Ui_MainWindow

from PyQt5.QtGui import QImage,QPixmap
from PyQt5.QtWidgets import QFileDialog

import cv2 as cv
import train_example

from Unet.train import *
from Unet.predict import *

import openpyxl 
import os
import glob

class GUI_Controller(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        
        
        self.save_RCNNmodel='E:/Nick/Desktop/SpineDataset/model_fasterrcnn_notation13.pkl'
        self.save_Unetmodel='E:/Nick/Desktop/SpineDataset/UnetTrain.pth'
        
        #self.save_RCNNdata_dir='E:/Nick/Desktop/SpineDataset/'
        self.UnetSegmentPath='E:/Nick/Desktop/SpineDataset/UnetPredict/'
        
        self.wb_new=openpyxl.Workbook()              #new a excel's file
        self.column=1                               #excel column initial
        self.sheet=self.wb_new.create_sheet("工作表1")
        
        self.ui=Ui_MainWindow()
        self.ui.setupUi(self)
        self.setup_control()
        
        
    def setup_control(self):
        
        self.ui.trainData12.clicked.connect(self.traindata12)
        self.ui.trainData23.clicked.connect(self.traindata23)
        self.ui.trainData13.clicked.connect(self.traindata13)
        self.ui.LoadImg.clicked.connect(self.loadImg)
        self.ui.Predict.clicked.connect(self.prediction)
        self.ui.loadmodel.clicked.connect(self.loadModel)
        return
    
    def traindata12(self):  
        
        root='E:/Nick/Desktop/SpineDataset/data/f01/'
        '''
        in root subdir we need 5 dir : image(use in SegmentBox & SpineDataset) ,notation2(use in SpineDataset)
        label(use in SegmentBox as ground truth) , segment(use in SegmentBox as result Segmentation) 
        segment_mask(use in SegmentBox as ground truth Segmentation)
        
        '''
        predict=False
        
        
        train_example.FasterRCNN(root,70,self.save_RCNNmodel,predict,Img=None)
        UnetTrain(root,70,self.save_Unetmodel,self.UnetSegmentPath)
        '''
        psudo function: FasterRCNN(root,num_batch,Img=None,self.save_RCNNmodel_dir,predict)
                        Unet(root,num_batch,self.save_Unetmodel_dir,predict)
                        
        
        
        
        
        '''
    def traindata23(self):
        root='E:/Nick/Desktop/SpineDataset/data/f02/'
        predict=False
        train_example.FasterRCNN(root,70,self.save_RCNNmodel,predict,Img=None)
        UnetTrain(root,70,self.save_Unetmodel,self.UnetSegmentPath)
        
    def traindata13(self):
        
        root='E:/Nick/Desktop/SpineDataset/data/f03/'
        predict=False
        train_example.FasterRCNN(root,70,self.save_RCNNmodel,predict,Img=None)
        UnetTrain(root,70,self.save_Unetmodel,self.UnetSegmentPath)
        
        
    def StoreTrainImageDC(self):
        root='E:/Nick/Desktop/SpineDataset/data/predict/'                        #fasterRCNN segment's Img dir
        img_dir01='E:/Nick/Desktop/SpineDataset/data/f01/'
        img=list(os.listdir(os.path.join(img_dir01,"image")))
        
        for i in img:
            py_files1 = glob.glob(root+'segment/*.png')                          #remove file from last time prediction
            py_files2 = glob.glob(root+'segment_mask/*.png')
            py_files3 = glob.glob(self.UnetSegmentPath+'/*.jpg')
            
            for py_file in py_files1:
                try:
                    os.remove(py_file)
                except OSError as e:
                    print(f"Error:{ e.strerror}")
            
            for py_file in py_files2:
                try:
                    os.remove(py_file)
                except OSError as e:
                    print(f"Error:{ e.strerror}")
            
            for py_file in py_files3:
                try:
                    os.remove(py_file)
                except OSError as e:
                    print(f"Error:{ e.strerror}")
        
        
            img_dir=img_dir01+'image/'+i
            root='E:/Nick/Desktop/SpineDataset/data/predict/'
            predict=True
            sortedbbox,img_result =train_example.FasterRCNN(root,1,self.save_RCNNmodel,predict,Img=img_dir)
            DC=UnetPredict(self.save_Unetmodel,root,self.UnetSegmentPath)       #Unet predict
            DC_float=[]
            average=0
        
            for i in range(20):                                                 #Unet each spine DC
                if i<len(DC):
                    
                    DC_float.append(DC[i].item())
                    average=average+DC[i].item()
                    self.sheet.cell(row=i+1, column=self.column, value=DC[i].item()) 
                else:
                    DC_float.append(0)
                    
            average=average/len(DC)
            self.ui.average.setText(str(average))
            self.sheet.cell(row=25, column=self.column, value=average) 
            self.column+=1
            
        self.wb_new.save('DC_Coefficient_test12.xlsx')
        
        
        
    
    def loadModel(self):
        
        filename,_=QFileDialog.getOpenFileName(self,'Open file','./') #Para2:title's string Para3:default open dir
        print(filename)
        self.save_Unetmodel=filename
    
    def loadImg(self):
        
        filename,_=QFileDialog.getOpenFileName(self,'Open file','./') #Para2:title's string Para3:default open dir
        print(filename)
        
        self.display_img(filename,1)
        
        self.img_dir=filename
        
    def display_img(self,img_dir,idx):                  #idx: position of image to show on QImage(Left、Middle、Right)
        
        if idx==1:
            img=cv.imread(img_dir)
            img=cv.resize(img, (250, 600), interpolation=cv.INTER_NEAREST)
            height,width,channel=img.shape
            bytesPerline=3*width
            qimg=QImage(img,width,height,bytesPerline,QImage.Format_RGB888).rgbSwapped()
        
            self.ui.InputIMG.setPixmap(QPixmap.fromImage(qimg))
            
        elif idx==2:                
            img=img_dir
            img=cv.resize(img, (250, 600), interpolation=cv.INTER_NEAREST)
            height,width,channel=img.shape
            bytesPerline=3*width
            qimg=QImage(img,width,height,bytesPerline,QImage.Format_RGB888).rgbSwapped()
            self.ui.ImgDetect.setPixmap(QPixmap.fromImage(qimg))
        elif idx==3:
            img=img_dir
            img=cv.resize(img, (250, 600), interpolation=cv.INTER_NEAREST)
            height,width=img.shape
            bytesPerline=width                      # 4 for RGBA, 3 for RGB so 1 for grayscale
            qimg=QImage(img,width,height,bytesPerline,QImage.Format_Indexed8)
            self.ui.SegmanetIMG.setPixmap(QPixmap.fromImage(qimg))
            
        
    def prediction(self):
        root='E:/Nick/Desktop/SpineDataset/data/predict/'
        
        '''
        In root sub_dir , we need 1 dir:segment
        '''
        


        py_files1 = glob.glob(root+'segment/*.png')                          #remove file from last time prediction
        py_files2 = glob.glob(root+'segment_mask/*.png')
        py_files3 = glob.glob(self.UnetSegmentPath+'/*.jpg')
        
        for py_file in py_files1:
            try:
                os.remove(py_file)
            except OSError as e:
                print(f"Error:{ e.strerror}")
        
        for py_file in py_files2:
            try:
                os.remove(py_file)
            except OSError as e:
                print(f"Error:{ e.strerror}")
        
        for py_file in py_files3:
            try:
                os.remove(py_file)
            except OSError as e:
                print(f"Error:{ e.strerror}")
        
        
        
        #FasterRCNN
        img_dir=self.img_dir
        predict=True
        sortedbbox,img_result =train_example.FasterRCNN(root,1,self.save_RCNNmodel,predict,Img=img_dir)
        self.display_img(img_result,2)
        self.ui.Detect_number.setText('Detected:'+str(len(sortedbbox)))
        
        #find contours
        label_dir=img_dir.replace("image", "label");                        #change dir from image to label
        label=cv.imread(label_dir)
        label=cv.cvtColor(label,cv.COLOR_BGR2GRAY)
        retval,thresh=cv.threshold(label,127,255,0)
        imgEdge,contours,hierarchy=cv.findContours(thresh,cv.RETR_LIST,cv.CHAIN_APPROX_SIMPLE)
        self.ui.number_GT.setText('GT:'+str(len(contours)))
        
        
        DC=UnetPredict(self.save_Unetmodel,root,self.UnetSegmentPath)       #Unet predict
        DC_float=[]
        average=0
        
        
        
        
        
        
        for i in range(20):                                                 #Unet each spine DC
            if i<len(DC):
                
                DC_float.append(DC[i].item())
                average=average+DC[i].item()
                self.sheet.cell(row=i+1, column=self.column, value=DC[i].item()) 
            else:
                DC_float.append(0)
                
            
        
        
        self.ui.V0.setText(str(DC_float[0]))
        self.ui.V1.setText(str(DC_float[1]))
        self.ui.V2.setText(str(DC_float[2]))
        self.ui.V3.setText(str(DC_float[3]))
        self.ui.V4.setText(str(DC_float[4]))
        self.ui.V5.setText(str(DC_float[5]))
        self.ui.V6.setText(str(DC_float[6]))
        self.ui.V7.setText(str(DC_float[7]))
        self.ui.V8.setText(str(DC_float[8]))
        self.ui.V9.setText(str(DC_float[9]))
        self.ui.V10.setText(str(DC_float[10]))
        self.ui.V11.setText(str(DC_float[11]))
        self.ui.V12.setText(str(DC_float[12]))
        self.ui.V13.setText(str(DC_float[13]))
        self.ui.V14.setText(str(DC_float[14]))
        self.ui.V15.setText(str(DC_float[15]))
        self.ui.V16.setText(str(DC_float[16]))
        self.ui.V17.setText(str(DC_float[17]))
        self.ui.V18.setText(str(DC_float[18]))
        self.ui.V19.setText(str(DC_float[19]))
        
        average=average/len(DC)
        self.ui.average.setText(str(average))
        self.sheet.cell(row=25, column=self.column, value=average) 
        self.column+=1
        
        img_result=cv.imread(self.img_dir,cv.IMREAD_GRAYSCALE)
        img=cv.imread(self.img_dir,cv.IMREAD_GRAYSCALE)
        
        height_width=img_result.shape                                  #let image to be black
        for h in range(img_result.shape[0]):
            for w in range(img_result.shape[1]):
                img_result[h,w]=0
        
        segment_file=list(os.listdir(self.UnetSegmentPath))
        
        
        
        for i in range(len(sortedbbox)):                                               #draw bbox's prediction
            predict=cv.imread(self.UnetSegmentPath+segment_file[i],cv.IMREAD_GRAYSCALE)
            for h in range(predict.shape[0]):
                for w in range(predict.shape[1]):
                    if predict[h,w]==255:
                        x=sortedbbox[i][0]
                        y=sortedbbox[i][1]
                        img_result[y+h,x+w]=255
                   
        dst = cv.addWeighted(img,0.7,img_result,0.3,0)              
                   
        self.display_img(dst,3)
        #cv.imshow('result',img)
        
        print('sortedbbox',sortedbbox)
        self.wb_new.save('DC_Coefficient.xlsx')
        #BBox_array=FasterRCNN(root,num_batch,Img=img,self.save_RCNNmodel_dir,,predict)
        
        #score,=UnetPredict(root,num_batch,self.save_Unetmodel_dir,self.save_RCNNdata_dir,self.UnetSegmentPath)
        
        
        
        
        
        
    
        
        
    
    
        
    